import openpyxl
from openpyxl.styles import PatternFill, Color

from func.task_list_func import *


@dp.message_handler(text='Задачи')
@logger.catch
async def push_task(msg: types.Message):
    logger.info(f'USER "{msg.from_user.id} - {await get_username_msg(msg)}" PUSH TASKS')
    user = await get_user(msg.from_user.id)
    if user:
        if len(user.participants) > 1:
            kb_chat = await create_show_chat_for_tasks_kb(user)
            mes = emojize('Выберите чат:')
            await bot.send_message(user.t_id, mes, reply_markup=kb_chat)
        elif len(user.participants) < 1:
            await msg.reply('Не однин ваш чат не зарегистрирован в "TaskBot"')
        else:
            kb_task = await create_show_tasks_kb(user.participants[0].chat, task_list=user.task_list)
            if user.super_admin:
                kb_task.insert(InlineKeyboardButton('Админы', callback_data=cb_admins.new(chat_id=user.participants[0].chat.chat_id)))
            mes = await create_show_tasks_mes(user.participants[0].chat, msg.from_user.id, task_list=user.task_list)
            await bot.send_message(msg.from_user.id, mes, reply_markup=kb_task)
    else:
        await msg.reply('Не однин ваш чат не зарегистрирован в "TaskBot"')


@dp.callback_query_handler(cb_chat_task.filter())
@logger.catch
async def choose_chat_for_task(call: types.CallbackQuery, callback_data: dict):
    user = await get_user(call.from_user.id)
    chat = await get_chat(callback_data['chat_id'])
    kb_task = await create_show_tasks_kb(chat, task_list=user.task_list)
    if user.super_admin:
        kb_task.insert(InlineKeyboardButton('Админы',
                       callback_data=cb_admins.new(chat_id=chat.chat_id)))
    mes = await create_show_tasks_mes(chat, call.from_user.id, task_list=user.task_list)
    await call.message.edit_text(emojize(mes), reply_markup=kb_task)
    await call.answer()


@dp.callback_query_handler(text='back_chat_task')
@logger.catch
async def back_chat_for_task(call: types.CallbackQuery):
    kb_chat = await create_show_chat_for_tasks_kb(await get_user(call.from_user.id))
    await call.message.edit_text('Выберите чат:', reply_markup=kb_chat)
    await call.answer()


@dp.callback_query_handler(cb_back_task.filter())
@logger.catch
async def back_task(call: types.CallbackQuery, callback_data: dict):
    user = await get_user(call.from_user.id)
    chat = await get_chat(callback_data['chat_id'])
    kb_task = await create_show_tasks_kb(chat, int(callback_data['page']), task_list=user.task_list)
    if user.super_admin:
        kb_task.insert(InlineKeyboardButton('Админы',
                       callback_data=cb_admins.new(chat_id=chat.chat_id)))
    mes = await create_show_tasks_mes(chat, call.from_user.id, int(callback_data['page']), task_list=user.task_list)
    await call.message.edit_text(mes, reply_markup=kb_task)
    await call.answer()


@dp.callback_query_handler(cb_page_list_task.filter())
@logger.catch
async def page_task(call: types.CallbackQuery, callback_data: dict):
    user = await get_user(call.from_user.id)
    chat = await get_chat(callback_data['chat_id'])
    kb_task = await create_show_tasks_kb(chat, page=int(callback_data['page']), task_list=user.task_list)
    if user.super_admin:
        kb_task.insert(InlineKeyboardButton('Админы',
                       callback_data=cb_admins.new(chat_id=chat.chat_id)))
    mes = await create_show_tasks_mes(chat, call.from_user.id, page=int(callback_data['page']), task_list=user.task_list)
    await call.message.edit_text(mes, reply_markup=kb_task)
    await call.answer()


@dp.callback_query_handler(cb_del_task.filter())
@logger.catch
async def del_task(call: types.CallbackQuery, callback_data: dict):
    user = await get_user(call.from_user.id)
    task = await get_task(callback_data['task_id'])
    chat, title = task.chat, task.title
    await del_task_by_id(callback_data['task_id'])
    kb_task = await create_show_tasks_kb(chat, task_list=user.task_list)
    if user.super_admin:
        kb_task.insert(InlineKeyboardButton('Админы',
                       callback_data=cb_admins.new(chat_id=chat.chat_id)))
    mes = emojize(f'Задача <b>{title}</b> удалена.\n')
    mes += await create_show_tasks_mes(chat, call.from_user.id, task_list=user.task_list)
    await call.message.edit_text(emojize(mes), reply_markup=kb_task)
    await call.answer()
    mes = f'<b>{user.name}</b> удалил задачу <u><b>{title}</b></u>'
    await bot.send_message(chat.chat_id, mes, reply_markup=kb_to_chat)


@dp.callback_query_handler(cb_week_tasks.filter())
@logger.catch
async def week_task(call: types.CallbackQuery, callback_data: dict):
    await update_task_list(call.from_user.id, 'week')
    await choose_chat_for_task(call=call, callback_data=callback_data)
    await call.answer()


@dp.callback_query_handler(cb_month_tasks.filter())
@logger.catch
async def month_task(call: types.CallbackQuery, callback_data: dict):
    await update_task_list(call.from_user.id, 'month')
    await choose_chat_for_task(call=call, callback_data=callback_data)
    await call.answer()


@dp.callback_query_handler(cb_all_tasks.filter())
@logger.catch
async def all_task(call: types.CallbackQuery, callback_data: dict):
    await update_task_list(call.from_user.id, 'all')
    await choose_chat_for_task(call=call, callback_data=callback_data)
    await call.answer()



@dp.callback_query_handler(cb_excel_tasks.filter())
@logger.catch
async def excel_task(call: types.CallbackQuery, callback_data: dict):
    list_column = ['Задача', 'Описание', 'Автор задачи', 'id', 'Выполнено', 'Дата создания', 'Срок выполнения',
                   'Дата выполнения']
    list_width = [20, 60, 20, 7, 15, 20, 20, 20]
    chat = await get_chat(callback_data['chat_id'])
    user = await get_user(call.from_user.id)
    wb = Workbook()
    if user.task_list == 'week':
        tasks_week, start, stop = await get_this_week_month_tasks(chat.chat_id, 'week')
        ws = wb.active
        file_name = f'{chat.title}_Задачи за неделю ({start.strftime("%d.%m.%Y")} - ' \
                    f'{(stop - datetime.timedelta(days=1)).strftime("%d.%m.%Y")}).xlsx'
        for n, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']):
            ws[f'{col}1'] = list_column[n]
            ws.column_dimensions[col].width = list_width[n]
        row = ws.row_dimensions[1]
        row.font = Font(bold=True, name='Calibri', size=12)
        for n, t in enumerate(tasks_week):
            user_create = await get_user(t.user_id_create)
            ws[f'A{n + 2}'] = t.title
            ws[f'B{n + 2}'] = t.description
            ws[f'C{n + 2}'] = user_create.name
            ws[f'D{n + 2}'] = t.id
            ws[f'E{n + 2}'] = 'Выполнено' if t.completed else 'Не выполнено'
            ws[f'F{n + 2}'] = t.date_create.strftime("%d.%m.%Y %H:%M") if t.date_create else ''
            ws[f'G{n + 2}'] = t.date_end.strftime("%d.%m.%Y") if t.date_end else ''
            ws[f'H{n + 2}'] = t.date_complete.strftime("%d.%m.%Y %H:%M") if t.date_complete else ''
            row = ws.row_dimensions[n + 2]
            if t.completed:
                row.fill = PatternFill("solid", fgColor="E2FFEC")
            else:
                row.fill = PatternFill("solid", fgColor="FFE2E2")
            bd = Side(style='thin', color="000000")
            row.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    elif user.task_list == 'month':
        tasks_month, start, stop = await get_this_week_month_tasks(chat.chat_id, 'month')
        ws = wb.active
        file_name = f'{chat.title}_Задачи за {start.strftime("%b %Y")}.xlsx'
        for n, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']):
            ws[f'{col}1'] = list_column[n]
            ws.column_dimensions[col].width = list_width[n]
        row = ws.row_dimensions[1]
        row.font = Font(bold=True, name='Calibri', size=12)
        for n, t in enumerate(tasks_month):
            user_create = await get_user(t.user_id_create)
            ws[f'A{n + 2}'] = t.title
            ws[f'B{n + 2}'] = t.description
            ws[f'C{n + 2}'] = user_create.name
            ws[f'D{n + 2}'] = t.id
            ws[f'E{n + 2}'] = 'Выполнено' if t.completed else 'Не выполнено'
            ws[f'F{n + 2}'] = t.date_create.strftime("%d.%m.%Y %H:%M") if t.date_create else ''
            ws[f'G{n + 2}'] = t.date_end.strftime("%d.%m.%Y") if t.date_end else ''
            ws[f'H{n + 2}'] = t.date_complete.strftime("%d.%m.%Y %H:%M") if t.date_complete else ''
            row = ws.row_dimensions[n + 2]
            if t.completed:
                row.fill = PatternFill("solid", fgColor="E2FFEC")
            else:
                row.fill = PatternFill("solid", fgColor="FFE2E2")
            bd = Side(style='thin', color="000000")
            row.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    else:
        file_name = f'{chat.title}_Задачи.xlsx'
        start_date = 0
        while True:
            oldest_task = await get_oldest_task(chat, start_date)
            if not oldest_task:
                break
            start, stop = await start_stop_week(oldest_task.date_end)
            ws = wb.create_sheet(f'{start.strftime("%d.%m.%Y")} - '
                                 f'{(stop - datetime.timedelta(days=1)).strftime("%d.%m.%Y")}')

            for n, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']):
                ws[f'{col}1'] = list_column[n]
                ws.column_dimensions[col].width = list_width[n]
            row = ws.row_dimensions[1]
            row.font = Font(bold=True, name='Calibri', size=12)

            tasks_week = await get_week_tasks(chat, start, stop)
            for n, t in enumerate(tasks_week):
                user_create = await get_user(t.user_id_create)
                ws[f'A{n + 2}'] = t.title
                ws[f'B{n + 2}'] = t.description
                ws[f'C{n + 2}'] = user_create.name
                ws[f'D{n + 2}'] = t.id
                ws[f'E{n + 2}'] = 'Выполнено' if t.completed else 'Не выполнено'
                ws[f'F{n + 2}'] = t.date_create.strftime("%d.%m.%Y %H:%M") if t.date_create else ''
                ws[f'G{n + 2}'] = t.date_end.strftime("%d.%m.%Y") if t.date_end else ''
                ws[f'H{n + 2}'] = t.date_complete.strftime("%d.%m.%Y %H:%M") if t.date_complete else ''
                row = ws.row_dimensions[n + 2]
                if t.completed:
                    row.fill = PatternFill("solid", fgColor="E2FFEC")
                else:
                    row.fill = PatternFill("solid", fgColor="FFE2E2")
                bd = Side(style='thin', color="000000")
                row.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            start_date = stop
        start, stop = await start_stop_date('week')
        try:
            wb.active = wb[f'{start.strftime("%d.%m.%Y")} - '
                                 f'{(stop - datetime.timedelta(days=1)).strftime("%d.%m.%Y")}']
        except KeyError:
            pass
    wb.save(file_name)
    await bot.send_document(call.from_user.id, open(file_name, 'rb'))
    os.remove(file_name)


    #
    #
    # user = await get_user(call.from_user.id)
    # task_pd = await get_excel_task(callback_data['chat_id'], user.task_list)
    # chat = await get_chat(callback_data['chat_id'])
    # if user.task_list == 'week':
    #     file_name = f'{chat.title}_Задачи за неделю.xlsx'
    # elif user.task_list == 'month':
    #     file_name = f'{chat.title}_Задачи за месяц.xlsx'
    # else:
    #     file_name = f'{chat.title}_Задачи.xlsx'
    # task_pd.to_excel(file_name)
    # wb = openpyxl.load_workbook(file_name)
    # sheet = wb.active
    # for col in sheet.columns:
    #     max_length = 0
    #     column = col[0].column_letter  # Get the column name
    #     for cell in col:
    #         try:  # Necessary to avoid error on empty cells
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(str(cell.value))
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     sheet.column_dimensions[column].width = adjusted_width
    # rows = sheet.max_row
    # cols = sheet.max_column
    # for i in range(1, rows + 1):
    #     if sheet.cell(row=i, column=5).value == 'Выполнено':
    #         for j in range(1, cols + 1):
    #             sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor="98FB98")
    #     else:
    #         for j in range(1, cols + 1):
    #             sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor="F08080")
    #
    # wb.save(file_name)
    # await bot.send_document(call.from_user.id, open(file_name, 'rb'))
    # await call.answer()
